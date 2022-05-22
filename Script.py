#  Libraries
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

#  path of the excel files

files_dir = Path.cwd() / "Excel_Files"

#  loop through folders and subfolders , read each file using pandas and append

#  each dataframe in a list variable
list_of_df = []
for path in list(files_dir.rglob("*.xls*")):
    df = pd.read_excel(path)
    list_of_df.append(df)

#  concatenate all the previous dataframes in one dataframe

df_all = pd.concat(list_of_df)

#  Create a new directory and save our previous dataframe on it
output_dir = Path.cwd() / "Output"
output_dir.mkdir(exist_ok=True)
df.to_excel(output_dir / "files_merged.xlsx",  index=False)

#  loop through all the xls file in files_dir directory , access to each file

#  using openpyxl and modify the first column name

#  save each file in a new directory

for path in list(files_dir.rglob("*.xls*")):
    wb = load_workbook(filename=path)
    ws = wb.worksheets[0]  # or wb["Sheet1"]
    ws["A1"] = "Key"
    output_with_new_column_name = Path.cwd() / "output_new_column_name"
    output_with_new_column_name.mkdir(exist_ok=True)
    wb.save(output_with_new_column_name / path.name)

#  loop through all the xls file in files_dir directory , access to each file

#  using Pandas and add new column which contains today date

#  save each file in a new directory and keeping the same directory structure

# as the original one

for path in list(files_dir.rglob("*.xls*")):
    df = pd.read_excel(path)
    df["new_column"] = pd.to_datetime("today")
    output_additional_column = Path.cwd() / "output_additional_column"
    output_additional_column.mkdir(exist_ok=True)
    (output_additional_column / path.parent.name).mkdir(exist_ok=True)
    df.to_excel(output_additional_column / path.parent.name / path.name,
                index=False)
